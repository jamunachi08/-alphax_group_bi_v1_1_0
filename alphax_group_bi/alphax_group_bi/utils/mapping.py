from __future__ import annotations
import csv
import io
import json
from typing import Dict, List

import frappe

def read_sample_map() -> List[Dict]:
    path = frappe.get_app_path("alphax_group_bi", "alphax_group_bi", "sample_maps", "example_map.csv")
    with open(path, "r", encoding="utf-8") as f:
        return list(csv.DictReader(f))

def template_rows_from_sample() -> List[Dict]:
    path = frappe.get_app_path("alphax_group_bi", "alphax_group_bi", "sample_maps", "example_map.json")
    with open(path, "r", encoding="utf-8") as f:
        payload = json.load(f)
    return payload.get("rows", [])

def _iter_excel_rows(ws):
    for row in ws.iter_rows(values_only=True):
        values = [("" if v is None else str(v).strip()) for v in row]
        if any(values):
            yield values

def parse_map_file(file_doc) -> List[Dict]:
    ext = (file_doc.file_name or "").lower().rsplit(".", 1)[-1]
    if ext == "csv":
        content = file_doc.get_content()
        if isinstance(content, bytes):
            content = content.decode("utf-8")
        return list(csv.DictReader(io.StringIO(content)))

    if ext in {"xlsx", "xlsm", "xltx", "xltm"}:
        try:
            import openpyxl
        except Exception:
            frappe.throw("openpyxl is required on the server to parse Excel uploads.")

        wb = openpyxl.load_workbook(file_doc.get_full_path(), data_only=True)
        sheet_name = "MAP" if "MAP" in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet_name]

        header = None
        output = []
        for values in _iter_excel_rows(ws):
            if header is None:
                # prefer the first row that contains classification-like headers
                joined = " | ".join(values).lower()
                if "classification" in joined or "chart of account" in joined:
                    header = [v or f"col_{i+1}" for i, v in enumerate(values)]
                    continue
                header = [v or f"col_{i+1}" for i, v in enumerate(values)]
                continue

            row = {}
            for i, value in enumerate(values):
                key = header[i] if i < len(header) else f"col_{i+1}"
                row[key] = value
            output.append(row)
        return output

    frappe.throw("Only CSV and XLSX files are supported for MAP import.")
